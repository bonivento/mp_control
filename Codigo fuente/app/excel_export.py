"""Exportación de resultados a Excel."""
from __future__ import annotations
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="005CAB", end_color="005CAB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(color="005CAB", bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row: int, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def construir_excel(estudio: dict, muestras: list[dict], resultados: dict) -> bytes:
    wb = Workbook()

    # Hoja 1: Trazabilidad
    ws = wb.active
    ws.title = "Trazabilidad"
    ws["A1"] = "Reporte de Control Estadístico de Calidad"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    info = [
        ("ID estudio", estudio.get("id")),
        ("Nombre", estudio.get("nombre")),
        ("Producto", estudio.get("producto")),
        ("Tipo", estudio.get("tipo")),
        ("Característica", estudio.get("caracteristica")),
        ("Unidad", estudio.get("unidad")),
        ("Analista", estudio.get("analista")),
        ("Lote", estudio.get("lote")),
        ("Tipo de gráfico", estudio.get("tipo_grafico")),
        ("LSL", estudio.get("lsl")),
        ("USL", estudio.get("usl")),
        ("Tamaño subgrupo", estudio.get("tamano_subgrupo")),
        ("Fecha creación", estudio.get("fecha_creacion")),
        ("Notas", estudio.get("notas")),
    ]
    for i, (k, v) in enumerate(info, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v if v is not None else "—")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40

    # Hoja 2: Datos
    ws2 = wb.create_sheet("Datos")
    ws2["A1"] = "Subgrupo"
    ws2["B1"] = "Fecha"
    ws2["C1"] = "Valores"
    _style_header(ws2, 1, 3)
    for i, m in enumerate(muestras, start=2):
        ws2.cell(row=i, column=1, value=m["subgrupo"])
        ws2.cell(row=i, column=2, value=m["fecha_muestra"])
        ws2.cell(row=i, column=3, value=json.dumps(m["valores"]))
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 60

    # Hoja 3: Resultados
    ws3 = wb.create_sheet("Resultados")
    ws3["A1"] = "Resultados del Análisis Estadístico"
    ws3["A1"].font = TITLE_FONT
    row = 3
    for key, value in (resultados or {}).items():
        ws3.cell(row=row, column=1, value=key).font = Font(bold=True)
        if isinstance(value, (dict, list)):
            ws3.cell(row=row, column=2, value=json.dumps(value, ensure_ascii=False, indent=2)[:32000])
        else:
            ws3.cell(row=row, column=2, value=value)
        row += 1
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
