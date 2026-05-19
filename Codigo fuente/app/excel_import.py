"""Importación de estudios desde archivos Excel (.xlsx).

Formato esperado:

Hoja 1 - "Trazabilidad" (key/value):
    A1: Campo            | B1: Valor
    A2: Nombre           | B2: ...
    A3: Producto         | B3: ...
    A4: Tipo             | B4: variable | atributo
    A5: Caracteristica   | B5: ...
    A6: Unidad           | B6: ...           (opcional)
    A7: Analista         | B7: ...           (opcional)
    A8: Lote             | B8: ...           (opcional)
    A9: Tipo de grafico  | B9: xr | xs | p | np | c | u
    A10: LSL             | B10: número       (opcional)
    A11: USL             | B11: número       (opcional)
    A12: Tamano subgrupo | B12: 2..25        (xr/xs)
    A13: Notas           | B13: ...          (opcional)

Hoja 2 - "Datos":
    Variables (xr/xs):
        Subgrupo | Med1 | Med2 | ... | MedN
        1        | val  | val  | ... | val
    Atributos p / np / u:
        Subgrupo | Defectivos | Tamano
    Atributos c:
        Subgrupo | Defectos
"""
from __future__ import annotations
import io
import unicodedata
from openpyxl import load_workbook


def _norm(s) -> str:
    if s is None:
        return ""
    txt = str(s).strip().lower()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return txt


_META_KEYS = {
    "nombre": "nombre",
    "producto": "producto",
    "tipo": "tipo",
    "caracteristica": "caracteristica",
    "característica": "caracteristica",
    "unidad": "unidad",
    "analista": "analista",
    "lote": "lote",
    "tipo de grafico": "tipo_grafico",
    "tipo de gráfico": "tipo_grafico",
    "tipo grafico": "tipo_grafico",
    "tipo_grafico": "tipo_grafico",
    "lsl": "lsl",
    "limite inferior": "lsl",
    "límite inferior": "lsl",
    "usl": "usl",
    "limite superior": "usl",
    "límite superior": "usl",
    "tamano subgrupo": "tamano_subgrupo",
    "tamaño subgrupo": "tamano_subgrupo",
    "tamano de subgrupo": "tamano_subgrupo",
    "tamaño de subgrupo": "tamano_subgrupo",
    "n": "tamano_subgrupo",
    "notas": "notas",
    "observaciones": "notas",
}


def _find_sheet(wb, candidates: list[str]):
    """Busca una hoja por nombre tolerando mayúsculas/acentos."""
    norm_map = {_norm(n): n for n in wb.sheetnames}
    for c in candidates:
        n = _norm(c)
        if n in norm_map:
            return wb[norm_map[n]]
    return None


def parse_excel(file_bytes: bytes) -> dict:
    """Parsea un archivo .xlsx y devuelve dict con metadata + muestras.

    Devuelve dict listo para pasar a `crear_estudio(payload)` + `muestras`:
        {
          "nombre": "...", "producto": "...", "tipo": "variable", ...,
          "muestras": [{"subgrupo": 1, "valores": [...]}, ...]
        }
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

    # 1) Trazabilidad
    sheet_meta = _find_sheet(wb, ["Trazabilidad", "Metadata", "Metadatos", "Info"])
    if not sheet_meta:
        raise ValueError(
            "No se encontró la hoja 'Trazabilidad'. Renombra la primera hoja "
            "a 'Trazabilidad' o usa el archivo de plantilla."
        )

    meta = {}
    for row in sheet_meta.iter_rows(values_only=True):
        if not row:
            continue
        key_raw = row[0] if len(row) > 0 else None
        val = row[1] if len(row) > 1 else None
        key_n = _norm(key_raw)
        target = _META_KEYS.get(key_n)
        if target:
            meta[target] = val

    required = ["nombre", "producto", "tipo", "caracteristica", "tipo_grafico"]
    missing = [r for r in required if not meta.get(r)]
    if missing:
        raise ValueError(
            "Faltan campos requeridos en la hoja Trazabilidad: " + ", ".join(missing)
        )

    tipo = _norm(meta["tipo"])
    if tipo not in ("variable", "atributo"):
        raise ValueError(
            f"Campo 'Tipo' inválido: '{meta['tipo']}'. Use 'variable' o 'atributo'."
        )
    meta["tipo"] = tipo

    grafico = _norm(meta["tipo_grafico"])
    if grafico not in ("xr", "xs", "p", "np", "c", "u"):
        raise ValueError(
            f"Tipo de gráfico inválido: '{meta['tipo_grafico']}'. "
            "Use xr, xs, p, np, c o u."
        )
    meta["tipo_grafico"] = grafico

    # Casts opcionales
    def _opt_float(k):
        v = meta.get(k)
        if v in (None, ""):
            meta[k] = None
        else:
            try:
                meta[k] = float(v)
            except (ValueError, TypeError):
                raise ValueError(f"Valor no numérico en '{k}': {v}")

    _opt_float("lsl")
    _opt_float("usl")

    n_val = meta.get("tamano_subgrupo")
    if n_val not in (None, ""):
        try:
            meta["tamano_subgrupo"] = int(n_val)
        except (ValueError, TypeError):
            raise ValueError(f"Tamaño de subgrupo no entero: {n_val}")
    else:
        meta["tamano_subgrupo"] = None

    # 2) Datos
    sheet_data = _find_sheet(wb, ["Datos", "Data", "Muestras"])
    if not sheet_data:
        raise ValueError(
            "No se encontró la hoja 'Datos'. Renombra la segunda hoja a 'Datos' "
            "o usa el archivo de plantilla."
        )

    rows = list(sheet_data.iter_rows(values_only=True))
    if not rows:
        raise ValueError("La hoja 'Datos' está vacía.")

    # Identificar la fila de encabezado: primera fila con texto en la col A no numérico
    header_idx = 0
    for i, r in enumerate(rows):
        if r and r[0] is not None and not isinstance(r[0], (int, float)):
            header_idx = i
            break
    data_rows = rows[header_idx + 1:]

    muestras = []
    if grafico in ("xr", "xs"):
        # Cada fila: subgrupo + mediciones
        for r in data_rows:
            if not r or r[0] is None:
                continue
            sub = int(r[0])
            valores = [float(v) for v in r[1:] if v is not None and v != ""]
            if not valores:
                continue
            muestras.append({"subgrupo": sub, "valores": valores})
        if not muestras:
            raise ValueError("No se leyeron filas de datos en la hoja 'Datos'.")
        n_detect = len(muestras[0]["valores"])
        for m in muestras:
            if len(m["valores"]) != n_detect:
                raise ValueError(
                    f"Subgrupo {m['subgrupo']} tiene {len(m['valores'])} mediciones; "
                    f"se esperaban {n_detect}."
                )
        if meta["tamano_subgrupo"] in (None, 0):
            meta["tamano_subgrupo"] = n_detect

    elif grafico in ("p", "u"):
        # Subgrupo | Defectivos/Defectos | Tamano
        for r in data_rows:
            if not r or r[0] is None:
                continue
            sub = int(r[0])
            if r[1] is None or r[2] is None:
                continue
            muestras.append({"subgrupo": sub, "valores": [int(r[1]), int(r[2])]})
    elif grafico == "np":
        # Subgrupo | Defectivos | Tamano (constante)
        for r in data_rows:
            if not r or r[0] is None or r[1] is None:
                continue
            sub = int(r[0])
            tam = int(r[2]) if len(r) > 2 and r[2] is not None else None
            if tam is None and muestras:
                tam = muestras[0]["valores"][1]
            if tam is None:
                raise ValueError(
                    "Para np debes incluir la columna 'Tamano muestra' (constante)."
                )
            muestras.append({"subgrupo": sub, "valores": [int(r[1]), tam]})
    elif grafico == "c":
        # Subgrupo | Defectos
        for r in data_rows:
            if not r or r[0] is None or r[1] is None:
                continue
            muestras.append({"subgrupo": int(r[0]), "valores": [int(r[1])]})

    if len(muestras) < 2:
        raise ValueError(
            f"Se necesitan al menos 2 subgrupos en la hoja 'Datos'; se leyeron {len(muestras)}."
        )

    meta["muestras"] = muestras
    # Eliminar campos vacíos opcionales que sean None
    return meta
