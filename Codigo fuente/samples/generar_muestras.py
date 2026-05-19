"""Genera archivos Excel de prueba para el sistema CEC.

Uso:
    python samples/generar_muestras.py

Crea los archivos en la misma carpeta. Cada archivo tiene dos hojas:
- Trazabilidad: metadata del estudio
- Datos: subgrupos / mediciones
"""
import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="005CAB", end_color="005CAB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
META_LABEL_FONT = Font(color="003A6B", bold=True)
TITLE_FONT = Font(color="005CAB", bold=True, size=14)
THIN = Side(style="thin", color="C5D2E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def _new_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _add_trazabilidad(wb, datos: list[tuple]):
    ws = wb.create_sheet("Trazabilidad")
    ws["A1"] = "Campo"
    ws["B1"] = "Valor"
    for c in ("A1", "B1"):
        ws[c].fill = HEADER_FILL
        ws[c].font = HEADER_FONT
        ws[c].alignment = Alignment(horizontal="center")
        ws[c].border = BORDER

    for i, (k, v) in enumerate(datos, start=2):
        ws.cell(row=i, column=1, value=k).font = META_LABEL_FONT
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=2, value=v).border = BORDER

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55
    return ws


def _add_data_header(ws, columnas: list[str]):
    for i, c in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = 14


# ============================================================
# 1) X-R en control - Peso de mango Tommy
# ============================================================
def gen_xr_en_control():
    random.seed(42)
    nombre_archivo = "1_xr_en_control_mango.xlsx"
    wb = _new_workbook()

    _add_trazabilidad(wb, [
        ("Nombre", "Peso de mango Tommy - En control"),
        ("Producto", "Mango Tommy"),
        ("Tipo", "variable"),
        ("Caracteristica", "Peso"),
        ("Unidad", "g"),
        ("Analista", "Estudiante demo"),
        ("Lote", "L-2026-001"),
        ("Tipo de grafico", "xr"),
        ("LSL", 235),
        ("USL", 265),
        ("Tamano subgrupo", 5),
        ("Notas", "Proceso estable, media 250 g, sigma within ~3 g"),
    ])

    ws = wb.create_sheet("Datos")
    _add_data_header(ws, ["Subgrupo", "Med 1", "Med 2", "Med 3", "Med 4", "Med 5"])
    media, sigma = 250.0, 3.0
    for sub in range(1, 26):
        ws.cell(row=sub + 1, column=1, value=sub).border = BORDER
        for j in range(5):
            val = round(media + random.gauss(0, sigma), 2)
            c = ws.cell(row=sub + 1, column=j + 2, value=val)
            c.border = BORDER
            c.number_format = "0.00"

    path = os.path.join(OUT_DIR, nombre_archivo)
    wb.save(path)
    return path


# ============================================================
# 2) X-R fuera de control - shift de proceso
# ============================================================
def gen_xr_fuera_control():
    random.seed(101)
    nombre_archivo = "2_xr_fuera_de_control_mango.xlsx"
    wb = _new_workbook()

    _add_trazabilidad(wb, [
        ("Nombre", "Peso de mango Tommy - Fuera de control"),
        ("Producto", "Mango Tommy"),
        ("Tipo", "variable"),
        ("Caracteristica", "Peso"),
        ("Unidad", "g"),
        ("Analista", "Estudiante demo"),
        ("Lote", "L-2026-002"),
        ("Tipo de grafico", "xr"),
        ("LSL", 235),
        ("USL", 265),
        ("Tamano subgrupo", 5),
        ("Notas", "Subgrupos 14-25 con shift +8 g (causa asignable)"),
    ])

    ws = wb.create_sheet("Datos")
    _add_data_header(ws, ["Subgrupo", "Med 1", "Med 2", "Med 3", "Med 4", "Med 5"])
    sigma = 3.0
    for sub in range(1, 26):
        ws.cell(row=sub + 1, column=1, value=sub).border = BORDER
        # Shift después del subgrupo 13: simula causa asignable (p. ej. cambio de proveedor)
        media = 250.0 if sub <= 13 else 258.0
        for j in range(5):
            val = round(media + random.gauss(0, sigma), 2)
            c = ws.cell(row=sub + 1, column=j + 2, value=val)
            c.border = BORDER
            c.number_format = "0.00"

    path = os.path.join(OUT_DIR, nombre_archivo)
    wb.save(path)
    return path


# ============================================================
# 3) X-S en control - Brix de sandía (n=10)
# ============================================================
def gen_xs_en_control():
    random.seed(7)
    nombre_archivo = "3_xs_en_control_brix_sandia.xlsx"
    wb = _new_workbook()

    _add_trazabilidad(wb, [
        ("Nombre", "Brix de sandía - En control"),
        ("Producto", "Sandía"),
        ("Tipo", "variable"),
        ("Caracteristica", "Grados Brix"),
        ("Unidad", "°Bx"),
        ("Analista", "Estudiante demo"),
        ("Lote", "L-2026-003"),
        ("Tipo de grafico", "xs"),
        ("LSL", 9.5),
        ("USL", 12.5),
        ("Tamano subgrupo", 10),
        ("Notas", "Proceso estable, media 11 °Bx, sigma within ~0.4"),
    ])

    ws = wb.create_sheet("Datos")
    headers = ["Subgrupo"] + [f"Med {i}" for i in range(1, 11)]
    _add_data_header(ws, headers)
    media, sigma = 11.0, 0.4
    for sub in range(1, 26):
        ws.cell(row=sub + 1, column=1, value=sub).border = BORDER
        for j in range(10):
            val = round(media + random.gauss(0, sigma), 2)
            c = ws.cell(row=sub + 1, column=j + 2, value=val)
            c.border = BORDER
            c.number_format = "0.00"

    path = os.path.join(OUT_DIR, nombre_archivo)
    wb.save(path)
    return path


# ============================================================
# 4) p en control - Aguacate con manchas
# ============================================================
def gen_p_en_control():
    random.seed(11)
    nombre_archivo = "4_p_en_control_aguacate.xlsx"
    wb = _new_workbook()

    _add_trazabilidad(wb, [
        ("Nombre", "Aguacate Hass con manchas - En control"),
        ("Producto", "Aguacate Hass"),
        ("Tipo", "atributo"),
        ("Caracteristica", "Frutos con manchas"),
        ("Analista", "Estudiante demo"),
        ("Lote", "L-2026-004"),
        ("Tipo de grafico", "p"),
        ("Notas", "Proporción estable ~5% defectuosos, n variable"),
    ])

    ws = wb.create_sheet("Datos")
    _add_data_header(ws, ["Subgrupo", "Defectivos", "Tamano muestra"])
    p_real = 0.05
    for sub in range(1, 26):
        n = random.randint(80, 120)
        # Simula binomial
        d = sum(1 for _ in range(n) if random.random() < p_real)
        ws.cell(row=sub + 1, column=1, value=sub).border = BORDER
        ws.cell(row=sub + 1, column=2, value=d).border = BORDER
        ws.cell(row=sub + 1, column=3, value=n).border = BORDER

    path = os.path.join(OUT_DIR, nombre_archivo)
    wb.save(path)
    return path


# ============================================================
# 5) p fuera de control - tendencia creciente de defectuosos
# ============================================================
def gen_p_fuera_control():
    random.seed(19)
    nombre_archivo = "5_p_fuera_de_control_aguacate.xlsx"
    wb = _new_workbook()

    _add_trazabilidad(wb, [
        ("Nombre", "Aguacate Hass con manchas - Fuera de control"),
        ("Producto", "Aguacate Hass"),
        ("Tipo", "atributo"),
        ("Caracteristica", "Frutos con manchas"),
        ("Analista", "Estudiante demo"),
        ("Lote", "L-2026-005"),
        ("Tipo de grafico", "p"),
        ("Notas", "Salto a 18% en subgrupos 16-25 (causa asignable)"),
    ])

    ws = wb.create_sheet("Datos")
    _add_data_header(ws, ["Subgrupo", "Defectivos", "Tamano muestra"])
    for sub in range(1, 26):
        n = 100
        p_real = 0.05 if sub <= 15 else 0.18
        d = sum(1 for _ in range(n) if random.random() < p_real)
        ws.cell(row=sub + 1, column=1, value=sub).border = BORDER
        ws.cell(row=sub + 1, column=2, value=d).border = BORDER
        ws.cell(row=sub + 1, column=3, value=n).border = BORDER

    path = os.path.join(OUT_DIR, nombre_archivo)
    wb.save(path)
    return path


# ============================================================
# 6) c fuera de control - defectos por lote de sábila
# ============================================================
def gen_c_fuera_control():
    random.seed(33)
    nombre_archivo = "6_c_fuera_de_control_sabila.xlsx"
    wb = _new_workbook()

    _add_trazabilidad(wb, [
        ("Nombre", "Defectos por lote de sábila - Tendencia"),
        ("Producto", "Sábila (Aloe vera)"),
        ("Tipo", "atributo"),
        ("Caracteristica", "Defectos visuales por lote"),
        ("Analista", "Estudiante demo"),
        ("Lote", "L-2026-006"),
        ("Tipo de grafico", "c"),
        ("Notas", "Tendencia creciente progresiva de defectos"),
    ])

    ws = wb.create_sheet("Datos")
    _add_data_header(ws, ["Subgrupo", "Defectos"])
    for sub in range(1, 26):
        # Media baja al inicio (3) que sube hasta ~12 al final
        media = 3 + (sub - 1) * 0.4
        defectos = max(0, int(round(media + random.gauss(0, 1.5))))
        ws.cell(row=sub + 1, column=1, value=sub).border = BORDER
        ws.cell(row=sub + 1, column=2, value=defectos).border = BORDER

    path = os.path.join(OUT_DIR, nombre_archivo)
    wb.save(path)
    return path


if __name__ == "__main__":
    paths = [
        gen_xr_en_control(),
        gen_xr_fuera_control(),
        gen_xs_en_control(),
        gen_p_en_control(),
        gen_p_fuera_control(),
        gen_c_fuera_control(),
    ]
    print("\nArchivos generados:")
    for p in paths:
        print("  •", os.path.basename(p))
