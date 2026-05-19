"""Exportación de resultados a Excel con hojas dedicadas y formato profesional."""
from __future__ import annotations
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList

from .tz import colombia_datetime, colombia_friendly


HEADER_FILL = PatternFill(start_color="005CAB", end_color="005CAB", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="005CAB", bold=True, size=14)
LABEL_FONT = Font(color="003A6B", bold=True)
THIN = Side(style="thin", color="C5D2E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Colores para estados (badge)
GREEN_FILL = PatternFill(start_color="D8F5D2", end_color="D8F5D2", fill_type="solid")
RED_FILL = PatternFill(start_color="FFE2E2", end_color="FFE2E2", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFF1DE", end_color="FFF1DE", fill_type="solid")


def _style_header(ws, row: int, n_cols: int, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _set_title(ws, text: str, row: int = 1, span: int = 6):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _fmt_num(v, digits=4):
    if v is None:
        return "—"
    if isinstance(v, (int, float)):
        if abs(v) >= 1e9 or (abs(v) < 1e-3 and v != 0):
            return f"{v:.{digits}e}"
        return round(float(v), digits)
    return v


def _hoja_trazabilidad(wb, estudio):
    ws = wb.active
    ws.title = "Trazabilidad"
    _set_title(ws, "Reporte de Control Estadístico de Calidad", row=1, span=2)
    info = [
        ("ID estudio", estudio.get("id")),
        ("Nombre", estudio.get("nombre")),
        ("Producto", estudio.get("producto")),
        ("Tipo", estudio.get("tipo")),
        ("Característica", estudio.get("caracteristica")),
        ("Unidad", estudio.get("unidad")),
        ("Analista", estudio.get("analista")),
        ("Lote", estudio.get("lote")),
        ("Tipo de gráfico", estudio.get("tipo_grafico")),
        ("LSL", estudio.get("lsl")),
        ("USL", estudio.get("usl")),
        ("Tamaño subgrupo", estudio.get("tamano_subgrupo")),
        ("Fecha creación (Colombia)", colombia_friendly(estudio.get("fecha_creacion"))),
        ("Notas", estudio.get("notas")),
    ]
    for i, (k, v) in enumerate(info, start=3):
        c1 = ws.cell(row=i, column=1, value=k)
        c1.font = LABEL_FONT
        c1.fill = SUBHEADER_FILL
        c1.border = BORDER
        c2 = ws.cell(row=i, column=2, value=v if v is not None else "—")
        c2.border = BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55


def _hoja_datos(wb, estudio, muestras):
    ws = wb.create_sheet("Datos")
    tipo = (estudio.get("tipo_grafico") or "").lower()
    # Cabeceras adaptadas al tipo
    if tipo in ("xr", "xs") and muestras:
        n = len(muestras[0]["valores"])
        headers = ["Subgrupo", "Fecha"] + [f"Med {i+1}" for i in range(n)]
    elif tipo in ("p", "np", "u"):
        headers = ["Subgrupo", "Fecha", "Defectivos/Defectos", "Tamaño muestra"]
    elif tipo == "c":
        headers = ["Subgrupo", "Fecha", "Defectos"]
    else:
        headers = ["Subgrupo", "Fecha", "Valores"]

    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for i, m in enumerate(muestras, start=2):
        ws.cell(row=i, column=1, value=m["subgrupo"]).border = BORDER
        ws.cell(row=i, column=2, value=colombia_datetime(m.get("fecha_muestra"))).border = BORDER
        vals = m["valores"]
        if tipo in ("xr", "xs"):
            for j, v in enumerate(vals):
                cell = ws.cell(row=i, column=3 + j, value=float(v))
                cell.number_format = "0.0000"
                cell.border = BORDER
        elif tipo in ("p", "np", "u"):
            ws.cell(row=i, column=3, value=vals[0]).border = BORDER
            ws.cell(row=i, column=4, value=vals[1] if len(vals) > 1 else "").border = BORDER
        elif tipo == "c":
            ws.cell(row=i, column=3, value=vals[0]).border = BORDER
        else:
            ws.cell(row=i, column=3, value=json.dumps(vals)).border = BORDER

    for c in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=c).column_letter
        ws.column_dimensions[col_letter].width = 14 if c <= 2 else 12


def _hoja_grafico_control(wb, resultados):
    chart_data = resultados.get("grafico_control")
    if not chart_data:
        return
    ws = wb.create_sheet("Gráfico de control")

    # Si trae x_chart y r/s_chart (variables), itera; si no es uno solo (atributos)
    subcharts = []
    if "x_chart" in chart_data:
        subcharts.append(("X̄ — Media", chart_data["x_chart"]))
        if "r_chart" in chart_data:
            subcharts.append(("R — Rango", chart_data["r_chart"]))
        if "s_chart" in chart_data:
            subcharts.append(("S — Desviación", chart_data["s_chart"]))
    else:
        subcharts.append((chart_data.get("title", "Gráfico de control"), chart_data))

    row = 1
    for titulo, ch in subcharts:
        _set_title(ws, titulo, row=row, span=6)
        row += 2

        ws.cell(row=row, column=1, value="Subgrupo")
        ws.cell(row=row, column=2, value="Valor")
        ws.cell(row=row, column=3, value="LC")
        ws.cell(row=row, column=4, value="LSC")
        ws.cell(row=row, column=5, value="LIC")
        ws.cell(row=row, column=6, value="Estado")
        _style_header(ws, row, 6)

        ucl = ch.get("ucl"); lcl = ch.get("lcl"); cl = ch.get("cl")
        out = set(ch.get("out_of_control") or [])
        data_start_row = row + 1
        for i, (sg, pt) in enumerate(zip(ch["subgroups"], ch["points"])):
            r = data_start_row + i
            ws.cell(row=r, column=1, value=sg).border = BORDER
            cell_pt = ws.cell(row=r, column=2, value=float(pt))
            cell_pt.number_format = "0.0000"; cell_pt.border = BORDER
            ws.cell(row=r, column=3, value=float(cl)).border = BORDER
            ws.cell(row=r, column=3).number_format = "0.0000"
            u_val = ucl[i] if isinstance(ucl, list) else ucl
            l_val = lcl[i] if isinstance(lcl, list) else lcl
            ws.cell(row=r, column=4, value=float(u_val)).border = BORDER
            ws.cell(row=r, column=4).number_format = "0.0000"
            ws.cell(row=r, column=5, value=float(l_val)).border = BORDER
            ws.cell(row=r, column=5).number_format = "0.0000"
            estado_cell = ws.cell(row=r, column=6,
                                  value=("FUERA DE CONTROL" if sg in out else "OK"))
            estado_cell.border = BORDER
            estado_cell.fill = RED_FILL if sg in out else GREEN_FILL
            estado_cell.alignment = Alignment(horizontal="center")

        data_end_row = data_start_row + len(ch["subgroups"]) - 1

        # Resumen
        row = data_end_row + 2
        ws.cell(row=row, column=1, value="Línea Central (LC)").font = LABEL_FONT
        ws.cell(row=row, column=2, value=_fmt_num(cl))
        row += 1
        ws.cell(row=row, column=1, value="LSC (promedio)").font = LABEL_FONT
        u_p = (sum(ucl)/len(ucl)) if isinstance(ucl, list) else ucl
        ws.cell(row=row, column=2, value=_fmt_num(u_p))
        row += 1
        ws.cell(row=row, column=1, value="LIC (promedio)").font = LABEL_FONT
        l_p = (sum(lcl)/len(lcl)) if isinstance(lcl, list) else lcl
        ws.cell(row=row, column=2, value=_fmt_num(l_p))
        row += 1
        ws.cell(row=row, column=1, value="Puntos fuera de control").font = LABEL_FONT
        ws.cell(row=row, column=2, value=len(ch.get("out_of_control") or []))
        row += 1
        ws.cell(row=row, column=1, value="Violaciones de reglas Nelson").font = LABEL_FONT
        ws.cell(row=row, column=2, value=len(ch.get("rules_violations") or []))
        row += 2

        # Gráfico nativo de Excel
        chart = LineChart()
        chart.title = titulo
        chart.style = 2
        chart.y_axis.title = "Valor"
        chart.x_axis.title = "Subgrupo"
        data_ref = Reference(ws, min_col=2, max_col=5,
                             min_row=data_start_row - 1, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 18
        chart.height = 9
        ws.add_chart(chart, f"H{data_start_row - 1}")

    for col, w in [("A", 12), ("B", 14), ("C", 12), ("D", 12), ("E", 12), ("F", 22)]:
        ws.column_dimensions[col].width = w


def _hoja_normalidad(wb, resultados, estudio):
    norm = resultados.get("normalidad")
    if not norm:
        return
    ws = wb.create_sheet("Normalidad")

    _set_title(ws, "Pruebas de normalidad — Datos del estudio", row=1, span=5)
    ws.cell(row=2, column=1,
            value=f"Aplicado a {norm.get('descriptive', {}).get('n', 0)} mediciones "
                  f"({estudio.get('caracteristica')} en {estudio.get('unidad') or '—'}). "
                  "α = 0.05.").font = Font(italic=True, color="5B6B7A")

    # --- Resultados de las pruebas ---
    row = 4
    ws.cell(row=row, column=1, value="Prueba")
    ws.cell(row=row, column=2, value="Estadístico")
    ws.cell(row=row, column=3, value="p-valor")
    ws.cell(row=row, column=4, value="¿Normal? (α=0.05)")
    ws.cell(row=row, column=5, value="Interpretación")
    _style_header(ws, row, 5)
    row += 1

    for key in ("shapiro", "anderson", "dagostino"):
        t = norm.get(key, {})
        if "error" in t:
            ws.cell(row=row, column=1, value=t.get("test", key)).border = BORDER
            for c in range(2, 5):
                ws.cell(row=row, column=c, value="—").border = BORDER
            ws.cell(row=row, column=5, value=t["error"]).border = BORDER
        else:
            ws.cell(row=row, column=1, value=t.get("test", key)).border = BORDER
            ws.cell(row=row, column=2, value=_fmt_num(t.get("statistic"))).border = BORDER
            ws.cell(row=row, column=3,
                    value=_fmt_num(t.get("p_value")) if "p_value" in t else "—").border = BORDER
            es_normal = t.get("normal", False)
            cell = ws.cell(row=row, column=4, value="Sí" if es_normal else "No")
            cell.fill = GREEN_FILL if es_normal else RED_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            ws.cell(row=row, column=5, value=t.get("interpretation", "")).border = BORDER
        row += 1

    # --- Veredicto consolidado ---
    row += 1
    tests = [norm.get(k, {}) for k in ("shapiro", "anderson", "dagostino")]
    valid = [t for t in tests if "error" not in t]
    if valid:
        normales = sum(1 for t in valid if t.get("normal"))
        n_total = len(valid)
        if normales == n_total:
            veredicto = (f"✓ Los datos PASAN las {n_total} pruebas de normalidad. "
                         "El análisis paramétrico (gráficos X̄-R/X̄-S, índices de capacidad) "
                         "es estadísticamente apropiado.")
            fill = GREEN_FILL
        elif normales == 0:
            veredicto = (f"✗ Los datos NO pasan ninguna de las {n_total} pruebas. "
                         "Considere transformar los datos (log, raíz, Box-Cox) o usar "
                         "técnicas no paramétricas antes de calcular Cp/Cpk.")
            fill = RED_FILL
        else:
            veredicto = (f"⚠ Resultado mixto: {normales}/{n_total} pruebas aceptan "
                         "normalidad. Los gráficos X̄-R y X̄-S son robustos ante "
                         "desviaciones leves; los índices de capacidad deben reportarse "
                         "con esta salvedad.")
            fill = ORANGE_FILL
        ws.cell(row=row, column=1, value="Veredicto consolidado").font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=veredicto)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 50
        row += 2

    # --- Estadística descriptiva ---
    desc = norm.get("descriptive", {})
    if desc:
        ws.cell(row=row, column=1, value="Estadística descriptiva de los datos").font = TITLE_FONT
        row += 2
        ws.cell(row=row, column=1, value="Métrica")
        ws.cell(row=row, column=2, value="Valor")
        _style_header(ws, row, 2)
        row += 1
        descs = [
            ("n", desc.get("n"), 0),
            ("Media", desc.get("mean"), 4),
            ("Mediana", desc.get("median"), 4),
            ("Desviación estándar", desc.get("std"), 4),
            ("Varianza", desc.get("var"), 4),
            ("Mínimo", desc.get("min"), 4),
            ("Máximo", desc.get("max"), 4),
            ("Rango", desc.get("range"), 4),
            ("Q1 (cuartil 1)", desc.get("q1"), 4),
            ("Q3 (cuartil 3)", desc.get("q3"), 4),
            ("Rango intercuartílico (IQR)", desc.get("iqr"), 4),
            ("Asimetría (skewness)", desc.get("skewness"), 4),
            ("Curtosis", desc.get("kurtosis"), 4),
            ("Coef. de variación (%)", desc.get("cv"), 2),
        ]
        for label, val, digits in descs:
            ws.cell(row=row, column=1, value=label).border = BORDER
            ws.cell(row=row, column=1).fill = SUBHEADER_FILL
            ws.cell(row=row, column=2, value=_fmt_num(val, digits)).border = BORDER
            row += 1

    # --- Histograma como datos + gráfico nativo ---
    hist = norm.get("histogram", {})
    if hist and hist.get("counts"):
        row += 1
        ws.cell(row=row, column=1, value="Histograma de frecuencias").font = TITLE_FONT
        row += 2
        ws.cell(row=row, column=1, value="Centro de clase")
        ws.cell(row=row, column=2, value="Frecuencia")
        _style_header(ws, row, 2)
        hist_start = row + 1
        for centro, count in zip(hist["centers"], hist["counts"]):
            row += 1
            ws.cell(row=row, column=1, value=float(centro)).number_format = "0.0000"
            ws.cell(row=row, column=2, value=int(count))
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2).border = BORDER
        hist_end = row

        # Gráfico
        chart = BarChart()
        chart.type = "col"
        chart.title = "Histograma de los datos"
        chart.style = 2
        chart.y_axis.title = "Frecuencia"
        chart.x_axis.title = "Valor"
        data_ref = Reference(ws, min_col=2, min_row=hist_start - 1, max_row=hist_end)
        cats_ref = Reference(ws, min_col=1, min_row=hist_start, max_row=hist_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 16; chart.height = 9
        ws.add_chart(chart, f"D{hist_start - 1}")

    # --- Q-Q plot ---
    qq = norm.get("qq_plot", {})
    if qq and qq.get("sample"):
        row += 3
        ws.cell(row=row, column=1, value="Gráfico Q-Q (cuantiles normales)").font = TITLE_FONT
        row += 2
        ws.cell(row=row, column=1, value="Cuantiles teóricos")
        ws.cell(row=row, column=2, value="Cuantiles muestrales")
        _style_header(ws, row, 2)
        qq_start = row + 1
        for x, y in zip(qq["theoretical"], qq["sample"]):
            row += 1
            ws.cell(row=row, column=1, value=float(x)).number_format = "0.0000"
            ws.cell(row=row, column=2, value=float(y)).number_format = "0.0000"
            ws.cell(row=row, column=1).border = BORDER
            ws.cell(row=row, column=2).border = BORDER
        qq_end = row

        chart = ScatterChart()
        chart.title = "Q-Q Plot"
        chart.style = 2
        chart.x_axis.title = "Cuantiles teóricos"
        chart.y_axis.title = "Cuantiles muestrales"
        xvals = Reference(ws, min_col=1, min_row=qq_start, max_row=qq_end)
        yvals = Reference(ws, min_col=2, min_row=qq_start, max_row=qq_end)
        series = Series(yvals, xvals, title="Datos")
        chart.series.append(series)
        chart.width = 16; chart.height = 9
        ws.add_chart(chart, f"D{qq_start - 1}")

    for col, w in [("A", 28), ("B", 18), ("C", 14), ("D", 14), ("E", 50)]:
        ws.column_dimensions[col].width = w


def _hoja_capacidad(wb, resultados):
    cap = resultados.get("capacidad")
    if not cap or "error" in cap:
        return
    ws = wb.create_sheet("Capacidad")
    _set_title(ws, "Índices de capacidad del proceso", row=1, span=3)
    row = 3
    ws.cell(row=row, column=1, value="Índice")
    ws.cell(row=row, column=2, value="Valor")
    ws.cell(row=row, column=3, value="Notas")
    _style_header(ws, row, 3)
    row += 1

    items = [
        ("Media (μ)", cap.get("mean"), "Tendencia central del proceso"),
        ("σ within (corto plazo)", cap.get("sigma_within"),
         "Estimación basada en variación dentro de subgrupos (R̄/d₂ o S̄/c₄)"),
        ("σ overall (largo plazo)", cap.get("sigma_overall"),
         "Desviación estándar muestral total"),
        ("LSL", cap.get("lsl"), "Límite inferior de especificación"),
        ("USL", cap.get("usl"), "Límite superior de especificación"),
        ("Cp", cap.get("Cp"), "Capacidad potencial (usa σ within)"),
        ("Cpk", cap.get("Cpk"), "Capacidad real (centrada, usa σ within)"),
        ("Pp", cap.get("Pp"), "Performance potencial (usa σ overall)"),
        ("Ppk", cap.get("Ppk"), "Performance real (usa σ overall)"),
        ("% fuera de especificación", cap.get("percent_out_of_spec"),
         "Porcentaje teórico bajo normalidad"),
        ("PPM fuera de especificación", cap.get("ppm_out_of_spec"),
         "Partes por millón teóricas"),
    ]
    for label, val, nota in items:
        if val is None:
            continue
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).fill = SUBHEADER_FILL
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=_fmt_num(val)).border = BORDER
        # Colorear Cpk según rango
        if label == "Cpk" and isinstance(val, (int, float)):
            if val >= 1.33:
                ws.cell(row=row, column=2).fill = GREEN_FILL
            elif val >= 1.0:
                ws.cell(row=row, column=2).fill = ORANGE_FILL
            else:
                ws.cell(row=row, column=2).fill = RED_FILL
        ws.cell(row=row, column=3, value=nota).border = BORDER
        row += 1

    if cap.get("interpretation"):
        row += 1
        ws.cell(row=row, column=1, value="Interpretación").font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=cap["interpretation"])
        cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 35

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 65


def _hoja_reglas(wb, resultados):
    chart_data = resultados.get("grafico_control") or {}
    violations = []
    if "x_chart" in chart_data:
        for v in chart_data["x_chart"].get("rules_violations") or []:
            violations.append({**v, "chart": "X̄"})
        for v in chart_data.get("r_chart", {}).get("rules_violations") or []:
            violations.append({**v, "chart": "R"})
        for v in chart_data.get("s_chart", {}).get("rules_violations") or []:
            violations.append({**v, "chart": "S"})
    else:
        for v in chart_data.get("rules_violations") or []:
            violations.append({**v, "chart": chart_data.get("type", "—")})

    if not violations:
        return

    ws = wb.create_sheet("Reglas de Nelson")
    _set_title(ws, "Violaciones a las reglas de Nelson", row=1, span=4)
    ws.cell(row=2, column=1,
            value="Detección automática de patrones no aleatorios en los datos del estudio."
            ).font = Font(italic=True, color="5B6B7A")
    row = 4
    ws.cell(row=row, column=1, value="Gráfico")
    ws.cell(row=row, column=2, value="Subgrupo")
    ws.cell(row=row, column=3, value="Regla #")
    ws.cell(row=row, column=4, value="Descripción")
    _style_header(ws, row, 4)
    for v in violations:
        row += 1
        ws.cell(row=row, column=1, value=v.get("chart", "—")).border = BORDER
        ws.cell(row=row, column=2, value=v.get("point")).border = BORDER
        ws.cell(row=row, column=3, value=v.get("rule")).border = BORDER
        ws.cell(row=row, column=4, value=v.get("desc")).border = BORDER

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 70


def construir_excel(estudio: dict, muestras: list[dict], resultados: dict) -> bytes:
    wb = Workbook()
    _hoja_trazabilidad(wb, estudio)
    _hoja_datos(wb, estudio, muestras)
    _hoja_grafico_control(wb, resultados or {})
    if (estudio.get("tipo") == "variable") or (resultados or {}).get("normalidad"):
        _hoja_normalidad(wb, resultados or {}, estudio)
    _hoja_capacidad(wb, resultados or {})
    _hoja_reglas(wb, resultados or {})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
